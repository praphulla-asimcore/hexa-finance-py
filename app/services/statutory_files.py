"""
Statutory submission file generators.
Amounts are taken EXACTLY from the CSI parsed data — no recalculation.
"""
import io
import base64
import hashlib
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _r2(n) -> float:
    return round(float(n or 0), 2)


def _month_label(yyyymm: str) -> str:
    names = ["January","February","March","April","May","June",
             "July","August","September","October","November","December"]
    if len(yyyymm) == 6:
        try:
            return f"{names[int(yyyymm[4:6])-1]} {yyyymm[:4]}"
        except Exception:
            pass
    return yyyymm


def _mmyyyy(yyyymm: str) -> str:
    """202606 → 06/2026"""
    return f"{yyyymm[4:6]}/{yyyymm[:4]}" if len(yyyymm) == 6 else yyyymm


def _ts() -> str:
    return datetime.utcnow().strftime("%Y%m%d%H%M%S")


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_info_block(ws, title: str, info_rows: list) -> int:
    """Write title + info rows, return next available row number."""
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].fill = PatternFill("solid", fgColor="4F46E5")
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    for i, (label, value) in enumerate(info_rows, 2):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=value).font = Font(size=10)
    return 2 + len(info_rows) + 1   # +1 blank row


def _write_col_headers(ws, row: int, headers: list) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="334155")
        c.alignment = Alignment(horizontal="center")


def _write_total_row(ws, row: int, values: list) -> None:
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="E2E8F0")


# ─── HRDF ─────────────────────────────────────────────────────────────────────

def generate_hrdf_file(submission: dict, employer_hrdf_code: str = "") -> dict:
    employees  = submission.get("employee_data", [])
    wage_month = submission.get("wage_month", "")
    cont_month = submission.get("contribution_month", "")
    entity     = submission.get("entity", "")
    ent_name   = submission.get("entity_name") or entity

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HRDF Levy"

    info = [
        ("Employer HRDF Code:", employer_hrdf_code or "—"),
        ("Entity:",             ent_name),
        ("Wage Month:",         _month_label(wage_month)),
        ("Contribution Month:", _month_label(cont_month)),
        ("Due Date:",           submission.get("due_date", "")),
    ]
    data_row = _write_info_block(ws, "HRDF Levy Contribution Schedule", info)

    headers = ["No.", "Employee Name", "IC / Passport No.", "Gross Salary (RM)", "HRDF Levy (RM)"]
    _write_col_headers(ws, data_row, headers)

    total_gross = total_levy = 0.0
    for i, emp in enumerate(employees, 1):
        gross = _r2(emp.get("grossSalary"))
        levy  = _r2(emp.get("hrdf"))
        total_gross += gross
        total_levy  += levy
        r = data_row + i
        for col, val in enumerate([i, emp.get("name",""), emp.get("idNumber",""), gross, levy], 1):
            ws.cell(row=r, column=col, value=val)

    total_r = data_row + len(employees) + 1
    _write_total_row(ws, total_r, ["", "TOTAL", "", round(total_gross, 2), round(total_levy, 2)])

    for col, w in [("A",6),("B",35),("C",22),("D",18),("E",16)]:
        ws.column_dimensions[col].width = w

    data = _save(wb)
    return {
        "file_data":      base64.b64encode(data).decode(),
        "file_name":      f"HRDF_{entity}_{wage_month}_{_ts()}.xlsx",
        "file_hash":      _sha256(data),
        "total_ee_amount": 0.0,
        "total_er_amount": round(total_levy, 2),
        "total_amount":    round(total_levy, 2),
    }


# ─── SOCSO + EIS ──────────────────────────────────────────────────────────────

def generate_socso_eis_file(submission: dict, employer_socso_no: str = "") -> dict:
    employees  = submission.get("employee_data", [])
    wage_month = submission.get("wage_month", "")
    cont_month = submission.get("contribution_month", "")
    entity     = submission.get("entity", "")
    ent_name   = submission.get("entity_name") or entity

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SOCSO + EIS"

    info = [
        ("Employer SOCSO No:", employer_socso_no or "—"),
        ("Entity:",            ent_name),
        ("Wage Month:",        _month_label(wage_month)),
        ("Contribution Month:",_month_label(cont_month)),
        ("Due Date:",          submission.get("due_date", "")),
    ]
    data_row = _write_info_block(ws, "SOCSO + EIS Contribution Schedule", info)

    headers = [
        "No.", "SOCSO No.", "Employee Name", "IC / Passport No.",
        "ee SOCSO (RM)", "er SOCSO (RM)", "ee EIS (RM)", "er EIS (RM)", "Total (RM)",
    ]
    _write_col_headers(ws, data_row, headers)

    t = dict(ee_s=0.0, er_s=0.0, ee_e=0.0, er_e=0.0)
    for i, emp in enumerate(employees, 1):
        ee_s = _r2(emp.get("socsoEmployee"))
        er_s = _r2(emp.get("socsoEmployer"))
        ee_e = _r2(emp.get("eisEmployee"))
        er_e = _r2(emp.get("eisEmployer"))
        tot  = ee_s + er_s + ee_e + er_e
        t["ee_s"] += ee_s; t["er_s"] += er_s
        t["ee_e"] += ee_e; t["er_e"] += er_e
        r = data_row + i
        for col, val in enumerate(
            [i, emp.get("socsoNumber",""), emp.get("name",""), emp.get("idNumber",""),
             ee_s, er_s, ee_e, er_e, round(tot,2)], 1):
            ws.cell(row=r, column=col, value=val)

    total_all = sum(t.values())
    total_r   = data_row + len(employees) + 1
    _write_total_row(ws, total_r, [
        "", "", "TOTAL", "",
        round(t["ee_s"],2), round(t["er_s"],2),
        round(t["ee_e"],2), round(t["er_e"],2),
        round(total_all, 2),
    ])

    for col, w in [("A",6),("B",14),("C",35),("D",22),
                   ("E",14),("F",14),("G",14),("H",14),("I",14)]:
        ws.column_dimensions[col].width = w

    data = _save(wb)
    return {
        "file_data":      base64.b64encode(data).decode(),
        "file_name":      f"SOCSO_EIS_{entity}_{wage_month}_{_ts()}.xlsx",
        "file_hash":      _sha256(data),
        "total_ee_amount": round(t["ee_s"] + t["ee_e"], 2),
        "total_er_amount": round(t["er_s"] + t["er_e"], 2),
        "total_amount":    round(total_all, 2),
    }


# ─── EPF — Borang A CSV format ────────────────────────────────────────────────

def _csv_escape(value: str) -> str:
    """Wrap in double-quotes if value contains comma, quote, or newline."""
    s = str(value)
    if any(c in s for c in (',', '"', '\n', '\r')):
        return '"' + s.replace('"', '""') + '"'
    return s


def generate_epf_file(submission: dict, employer_epf_no: str = "") -> dict:
    employees  = submission.get("employee_data", [])
    wage_month = submission.get("wage_month", "")
    entity     = submission.get("entity", "")

    valid = [e for e in employees
             if _r2(e.get("epfEmployee")) + _r2(e.get("epfEmployer")) > 0]

    total_ee = total_er = 0.0
    rows = ["Member EPF No,Employee Identification No,Employee Name,Employee Salary,Employer Amount,Employee Amount"]
    for emp in valid:
        ee    = _r2(emp.get("epfEmployee"))
        er    = _r2(emp.get("epfEmployer"))
        total_ee += ee
        total_er += er
        epf_no = _csv_escape((emp.get("epfNumber") or "").strip())
        ic     = _csv_escape((emp.get("idNumber")  or "").strip())
        name   = _csv_escape((emp.get("name")       or "").strip())
        salary = f"{_r2(emp.get('grossSalary')):.2f}"
        rows.append(f"{epf_no},{ic},{name},{salary},{er:.2f},{ee:.2f}")

    data = "\r\n".join(rows).encode("utf-8")
    return {
        "file_data":      base64.b64encode(data).decode(),
        "file_name":      f"EPF_{entity}_{wage_month}_{_ts()}.csv",
        "file_hash":      _sha256(data),
        "total_ee_amount": round(total_ee, 2),
        "total_er_amount": round(total_er, 2),
        "total_amount":    round(total_ee + total_er, 2),
    }


# ─── MTD / PCB — LHDNM e-Data PCB fixed-width text format ───────────────────
# Format spec: Manual Muatnaik e-Data PCB (LHDNM), Appendix 1
# Header: H + No.E Ibu Pejabat(10) + No.E Cawangan(10) + Year(4) + Month(2)
#           + Amaun PCB in cents(10) + Bil.PCB(5) + Amaun CP38 in cents(10) + Bil.CP38(5)
# Detail: D + TaxRef(11) + Name(60) + OldIC(12) + NewIC(12) + Passport(12)
#           + CountryCode(2) + PCB cents(8) + CP38 cents(8) + EmpNo(10)
# Amounts are in cents (integer), zero-padded left.

def _cents(amount) -> int:
    return round(_r2(amount) * 100)


def _strip_ic(ic: str) -> str:
    return ic.replace("-", "").replace(" ", "")


def generate_mtd_file(submission: dict, employer_mtd_no: str = "") -> dict:
    employees  = submission.get("employee_data", [])
    wage_month = submission.get("wage_month", "")
    entity     = submission.get("entity", "")

    year  = wage_month[:4]
    month = wage_month[4:6] if len(wage_month) >= 6 else "00"

    # Employer TIN: strip letter prefix, zero-pad left to 10 digits
    e_no = "".join(c for c in (employer_mtd_no or "") if c.isdigit()).zfill(10)[:10]

    valid = [e for e in employees if _r2(e.get("mtd")) > 0]

    total_pcb_cents  = sum(_cents(e.get("mtd"))      for e in valid)
    total_cp38_cents = sum(_cents(e.get("cp38", 0))  for e in valid)
    count_pcb        = len(valid)
    count_cp38       = sum(1 for e in valid if _cents(e.get("cp38", 0)) > 0)

    header = (
        "H"
        + e_no                                      # No. E Ibu Pejabat  (2-11)
        + e_no                                      # No. E Cawangan     (12-21)
        + year                                      # Tahun              (22-25)
        + month.zfill(2)                            # Bulan              (26-27)
        + str(total_pcb_cents).zfill(10)[:10]       # Amaun PCB          (28-37)
        + str(count_pcb).zfill(5)[:5]              # Bil. PCB           (38-42)
        + str(total_cp38_cents).zfill(10)[:10]      # Amaun CP38         (43-52)
        + str(count_cp38).zfill(5)[:5]             # Bil. CP38          (53-57)
    )

    lines = [header]
    for emp in valid:
        pcb_c  = str(_cents(emp.get("mtd"))).zfill(8)[:8]
        cp38_c = str(_cents(emp.get("cp38", 0))).zfill(8)[:8]

        tax_ref = (emp.get("taxRefNumber") or "").strip()[:11].ljust(11)
        name    = (emp.get("name") or "").upper()[:60].ljust(60)

        raw_id   = _strip_ic(emp.get("idNumber") or "")
        id_type  = (emp.get("idType") or "IC").upper()
        is_local = (emp.get("category") or "Local") != "Foreign"

        if is_local and "PASSPORT" not in id_type:
            kp_lama  = " " * 12
            kp_baru  = raw_id[:12].ljust(12)
            passport = " " * 12
        else:
            kp_lama  = " " * 12
            kp_baru  = " " * 12
            passport = raw_id[:12].ljust(12)

        nat = (emp.get("nationality") or "").upper()
        country = "MY" if (is_local or "MALAYSIA" in nat or nat == "MY") else "  "

        emp_no = (emp.get("employeeId") or "").strip()[:10].ljust(10)

        lines.append(
            "D"
            + tax_ref    # No. Rujukan Cukai  (2-12)
            + name       # Nama               (13-72)
            + kp_lama    # No. KP Lama        (73-84)
            + kp_baru    # No. KP Baru        (85-96)
            + passport   # No. Passport       (97-108)
            + country    # Kod Negara         (109-110)
            + pcb_c      # Amaun PCB          (111-118)
            + cp38_c     # Amaun CP38         (119-126)
            + emp_no     # No. Pekerja        (127-136)
        )

    data = "\r\n".join(lines).encode("utf-8")
    total_pcb = _r2(total_pcb_cents / 100)
    return {
        "file_data":      base64.b64encode(data).decode(),
        "file_name":      f"MTD_{entity}_{wage_month}_{_ts()}.txt",
        "file_hash":      _sha256(data),
        "total_ee_amount": total_pcb,
        "total_er_amount": 0.0,
        "total_amount":    total_pcb,
    }
