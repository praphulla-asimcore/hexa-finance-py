import io
from datetime import datetime, date
from openpyxl import load_workbook
import xlrd

from app.services.statutory_rates import (
    socso_contribution, eis_contribution, epf_contribution, epf_basis,
)


def _json_safe(obj):
    """Recursively convert non-JSON-serializable values (datetime, date) to strings."""
    if isinstance(obj, (datetime, date)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(v) for v in obj]
    return obj


# ─── CSI parser ───────────────────────────────────────────────────────────────

# New template column names (HSSB_CSI Template_New.xlsx)
# Old template column names kept as fallbacks for backwards compatibility.

_ENTITY_ALIASES: dict[str, str] = {
    "HEDU": "KISB",
}


def _to_num(val) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(str(val).replace(",", ""))
    except (ValueError, TypeError):
        return 0.0


def _get(row, col_map, *keys) -> float:
    """Return numeric value for the first key found in col_map."""
    for key in keys:
        idx = col_map.get(key)
        if idx is not None and idx < len(row):
            return _to_num(row[idx])
    return 0.0


def _get_str(row, col_map, *keys) -> str:
    """Return string value for the first key found in col_map."""
    for key in keys:
        idx = col_map.get(key)
        if idx is not None and idx < len(row):
            val = row[idx]
            if val is not None:
                return str(val).strip()
    return ""


def _build_col_map(header_row) -> dict:
    return {str(cell).strip(): idx for idx, cell in enumerate(header_row) if cell is not None}


def _parse_employee_my(row, col_map) -> dict | None:
    """Parse one CSI employee row against the Malaysian template. Returns
    None if row should be skipped."""
    emp_id_idx = col_map.get("Employee ID")
    if emp_id_idx is None or emp_id_idx >= len(row):
        return None
    emp_id_val = row[emp_id_idx]
    if emp_id_val is None or str(emp_id_val).strip() == "":
        return None

    ctc_hexa = _get(row, col_map, "CTC Hexa")
    if ctc_hexa == 0:
        return None

    return {
        "employeeId":    str(emp_id_val).strip(),
        # New template uses "Name"; old uses "Nickname / Name"
        "name":          _get_str(row, col_map, "Nickname / Name", "Name"),
        "nationality":   _get_str(row, col_map, "Nationality"),
        # New template uses "Client"; old uses "Cost Centre"
        "costCentre":    _get_str(row, col_map, "Cost Centre", "Client"),
        # Maybank CMS Favourite Beneficiary/Biller Code — written into col 4 of the
        # RCGEN2 bank file. Optional in the CSI; when present it overrides the
        # consultant-DB (Airtable) value during bank-file generation.
        "favouriteBeneficiaryCode": _get_str(
            row, col_map,
            "Favourite Beneficiary Code", "Favorite Beneficiary Code",
            "Favourite Beneficiary/Biller Code", "Favorite Beneficiary/Biller Code"),
        "clientType":    _get_str(row, col_map, "Client Type", "Margin Type").upper() or "CC",
        "grossSalary":   _get(row, col_map, "Gross Salary"),
        "basicSalary":   _get(row, col_map, "Basic Pay", "Basic Salary"),
        # New template uses "Claims Amount"; old uses "Claim"
        "claim":         _get(row, col_map, "Claim", "Claims Amount"),
        "bonus":         sum(_get(row, col_map, c) for c in [
                             "Commission (Daily/Weekly/Monthly)",
                             "Retirement Bonus (Monthly)",
                             "Retirement contribution"]),
        "caDedn":        sum(_get(row, col_map, c) for c in [
                             "Deduction", "Deduction (from Net Salary)"]),
        "epfEmployee":   _get(row, col_map, "EPF Employee"),
        "epfEmployer":   _get(row, col_map, "EPF Employer"),
        "eisEmployee":   _get(row, col_map, "EIS Employee"),
        "eisEmployer":   _get(row, col_map, "EIS Employer"),
        "socsoEmployee": _get(row, col_map, "SOCSO Employee"),
        "socsoEmployer": _get(row, col_map, "SOCSO Employer"),
        "hrdf":          _get(row, col_map, "HRDF"),
        "mtd":           _get(row, col_map, "MTD"),
        "ctcHexa":       ctc_hexa,
        # File's own CTC Hexa, preserved verbatim. statutory_enrich overwrites
        # ``ctcHexa`` with a rate-table recompute (for accrual/statutory use),
        # so keep the file value for margin checks that must compare against the
        # file's CTC Client on the SAME costing basis.
        "ctcHexaFile":   ctc_hexa,
        "netSalary":     _get(row, col_map, "Net Salary"),
        "ctcClient":     _get(row, col_map, "CTC Client"),
        "totalBilling":  _get(row, col_map, "Total Billing"),
        "mgmtFee":       _get(row, col_map, "Mgmt Fee (RM)"),
    }


def _parse_employee_id(row, col_map) -> dict | None:
    raise NotImplementedError(
        "Indonesia (PTHIT) CSI parsing is not yet configured — no column "
        "schema has been defined for this entity's payroll file format yet. "
        "A sample CSI file is needed to map its columns before this entity "
        "can be uploaded."
    )


def _parse_employee_np(row, col_map) -> dict | None:
    """Parse one CSI employee row against Nepal's payroll register format
    (mapped from Nepal_CSI_Sample.xlsx, 2026-08-03 — SBC/Time Energy/GW
    client sheets, one row per employee, no "Entity" column so each sheet is
    one client, same as Malaysia's older single-sheet-per-entity format).

    EIN (Employee Identification Number) is the canonical identifier,
    confirmed over EID (a per-file internal row number) — only EIN is stable
    enough to eventually match nepal_employee_master.employee_ledger_code.
    """
    ein_idx = col_map.get("EIN")
    if ein_idx is None or ein_idx >= len(row):
        return None
    ein_val = row[ein_idx]
    if ein_val is None or str(ein_val).strip() == "":
        return None

    gross = _get(row, col_map, "Gross Income")
    if gross == 0:
        return None

    # Some sheets (e.g. GW) carry a post-adjustment final payable figure
    # distinct from "Net Salary" -- prefer it only when that column actually
    # exists on this sheet, not just when its value happens to be non-zero.
    if "Net Payable After tax adjustment" in col_map:
        net_salary = _get(row, col_map, "Net Payable After tax adjustment")
    else:
        net_salary = _get(row, col_map, "Net Salary")

    ssf_contribution = _get(row, col_map, "SSF Contribution")  # employer-side, per file
    ssf_deduction    = _get(row, col_map, "SSF Deduction")     # employee-side, deducted from gross

    return {
        "employeeId":  str(ein_val).strip(),
        "eid":         _get_str(row, col_map, "EID"),
        "name":        _get_str(row, col_map, "Employee"),
        "costCentre":  _get_str(row, col_map, "Branch"),
        "designation": _get_str(row, col_map, "Designation"),
        "nationality": "Nepali",
        "basicSalary": _get(row, col_map, "Basic Salary"),
        "grossSalary": gross,
        "netSalary":   net_salary,
        # Nepal-specific statutory figures, extracted as-is from the file --
        # not recomputed (see app.services.statutory.np, which currently
        # passes these through unchanged pending confirmation of Nepal's
        # SSF-vs-PF scheme and TDS slabs).
        "ssfContribution": ssf_contribution,
        "ssfDeduction":    ssf_deduction,
        "cit":             _get(row, col_map, "CIT"),
        "sst":             _get(row, col_map, "SST"),
        "tds":             _get(row, col_map, "TDS"),
        "deductionPF":     _get(row, col_map, "Deduction PF"),
        "incomePF":        _get(row, col_map, "Income PF"),
        "totalDeduction":  _get(row, col_map, "Total Deduction"),
        # No confirmed Nepal margin/billing model yet -- this file has no
        # CTC Client/Total Billing/Mgmt Fee columns like Malaysia's CSI does.
        # ctcHexa is structurally required downstream (totalCTC summation),
        # so use gross + the file's own employer-side SSF figure as an
        # honest interim value (both real, present numbers -- nothing
        # fabricated) rather than 0, which would misleadingly read as "no
        # cost." Revisit once Nepal's margin/billing model is confirmed.
        "ctcHexa":     round(gross + ssf_contribution, 2),
        "ctcHexaFile": round(gross + ssf_contribution, 2),
        # Malaysia-specific concepts with no Nepal equivalent yet -- zeroed/
        # blank rather than guessed, since this file carries no billing data.
        "ctcClient":     0,
        "totalBilling":  0,
        "mgmtFee":       0,
        "favouriteBeneficiaryCode": "",
        "clientType":    "",
        "claim": 0, "bonus": 0, "caDedn": 0,
        "epfEmployee": 0, "epfEmployer": 0,
        "eisEmployee": 0, "eisEmployer": 0,
        "socsoEmployee": 0, "socsoEmployer": 0,
        "hrdf": 0, "mtd": 0,
    }


# Per-country CSI row parser, selected by the case's entity → country (see
# app.config.get_entity_country). Fail loud on an unrecognised/unconfigured
# country rather than silently parsing a foreign file against the Malaysian
# schema — a country-CSI mismatch could otherwise produce plausible-looking
# but wrong numbers instead of an obvious error.
CSI_PARSERS: dict = {
    "MY": _parse_employee_my,
    "ID": _parse_employee_id,
    "NP": _parse_employee_np,
}


def _get_csi_parser(country: str):
    parser = CSI_PARSERS.get((country or "MY").upper())
    if parser is None:
        raise NotImplementedError(
            f"No CSI parser configured for country {country!r}. "
            f"Supported: {sorted(CSI_PARSERS)}."
        )
    return parser


# Columns whose absence gets surfaced to the user as "missingColumns" on the
# entity — country-specific since Nepal's file never has Malaysia's literal
# "Employee ID"/"CTC Hexa" column names (and vice versa).
REQUIRED_COLUMNS: dict = {
    "MY": ["Employee ID", "CTC Hexa", "Gross Salary", "Net Salary"],
    "NP": ["EIN", "Employee", "Gross Income", "Net Salary"],
}


def _process_sheets(sheets: list[tuple[str, list]], country: str = "MY") -> list[dict]:
    """Build entities list from (sheet_name, rows) pairs."""
    parse_employee = _get_csi_parser(country)
    entities = []
    for sheet_name, rows in sheets:
        if not rows or len(rows) < 2:
            continue

        # Find header row — first row that contains "Employee ID"
        header_idx = 0
        for i, row in enumerate(rows):
            if row and any(str(c or "").strip() == "Employee ID" for c in row):
                header_idx = i
                break

        col_map = _build_col_map(rows[header_idx])

        # New template format: has an "Entity" column → group rows by entity value
        has_entity_col = "Entity" in col_map
        entity_col_idx = col_map.get("Entity")

        # Old format required cols; new format has different names
        required = REQUIRED_COLUMNS.get((country or "MY").upper(), REQUIRED_COLUMNS["MY"])
        missing_cols = [c for c in required if c not in col_map]

        if has_entity_col:
            # Group employees by the Entity column value
            from collections import defaultdict
            groups: dict[str, list] = defaultdict(list)
            for row in rows[header_idx + 1:]:
                if not row:
                    continue
                emp = parse_employee(row, col_map)
                if emp is None:
                    continue
                entity_val = str(row[entity_col_idx] or "").strip() if entity_col_idx < len(row) else ""
                group_key = _ENTITY_ALIASES.get(entity_val, entity_val) or sheet_name.strip()
                groups[group_key].append(emp)

            for ent_name, emps in groups.items():
                if emps:
                    entities.append({
                        "sheetName":      ent_name,
                        "employees":      emps,
                        "totalCTC":       round(sum(e["ctcHexa"] for e in emps), 2),
                        "missingColumns": missing_cols,
                    })
        else:
            # Old format: one entity per sheet
            employees = []
            for row in rows[header_idx + 1:]:
                if not row:
                    continue
                emp = parse_employee(row, col_map)
                if emp:
                    employees.append(emp)

            if employees:
                entities.append({
                    "sheetName":      sheet_name.strip(),
                    "employees":      employees,
                    "totalCTC":       round(sum(e["ctcHexa"] for e in employees), 2),
                    "missingColumns": missing_cols,
                })

    return entities


def parse_excel_buffer(data: bytes, country: str = "MY") -> list[dict]:
    # Try openpyxl first (handles .xlsx / .xlsm)
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = [
            (name, list(wb[name].iter_rows(values_only=True)))
            for name in wb.sheetnames
        ]
        wb.close()
        return _json_safe(_process_sheets(sheets, country))
    except Exception as e:
        if "zip" not in str(e).lower() and "openpyxl" not in str(e).lower():
            raise  # unexpected error — re-raise

    # Fall back to xlrd for legacy .xls (Excel 97-2003 binary format)
    try:
        wb = xlrd.open_workbook(file_contents=data)
        sheets = [
            (ws.name, [ws.row_values(r) for r in range(ws.nrows)])
            for ws in wb.sheets()
        ]
        return _json_safe(_process_sheets(sheets, country))
    except Exception as e:
        raise ValueError(
            f"Could not read file as .xlsx or .xls: {e}. "
            "Please save as Excel Workbook (.xlsx) and re-upload."
        )


# ─── Payroll parser ───────────────────────────────────────────────────────────
# Handles the HSSB 37-column payroll report format.
# Header row: Employee ID | Employee Name | ... | Net Pay | rEPF | ... | Bank Account Number
# Rows 0–2 are metadata; header is the first row where col 0 == "Employee ID".

_PAYROLL_REQUIRED = ["Employee ID", "Employee Name", "Net Pay"]

_SKIP_PREFIXES = ("total", "recap", "employee statutory", "employer statutory")


def _is_payroll_summary_row(emp_id_str: str) -> bool:
    low = emp_id_str.lower()
    return any(low.startswith(p) for p in _SKIP_PREFIXES)


def parse_payroll_excel_buffer(data: bytes) -> list[dict]:
    """Parse the HSSB payroll report (37-column, multi-row header format)."""
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        all_rows = list(wb.active.iter_rows(values_only=True))
        company_name = ""
        if all_rows and all_rows[0] and all_rows[0][1]:
            company_name = str(all_rows[0][1]).strip()
        wb.close()
    except Exception as e:
        raise ValueError(f"Could not read payroll file: {e}")

    # Locate header row: first row where col 0 == "Employee ID"
    header_idx = None
    for i, row in enumerate(all_rows):
        if row and str(row[0] or "").strip().lower() == "employee id":
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(
            "No 'Employee ID' header found. "
            "Check that row 4 of the payroll file contains the column headers."
        )

    header = all_rows[header_idx]
    col_map: dict[str, int] = {str(c or "").strip(): idx for idx, c in enumerate(header) if c}
    missing_cols = [c for c in _PAYROLL_REQUIRED if c not in col_map]

    def _pv(row: tuple, key: str) -> float:
        idx = col_map.get(key)
        return _to_num(row[idx]) if idx is not None and idx < len(row) else 0.0

    def _sv(row: tuple, key: str) -> str:
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        return str(row[idx] or "").strip()

    employees = []
    for row in all_rows[header_idx + 1:]:
        if not row:
            continue
        raw_id = row[col_map["Employee ID"]] if "Employee ID" in col_map and col_map["Employee ID"] < len(row) else None
        if raw_id is None:
            continue
        emp_id_str = str(raw_id).strip()
        if not emp_id_str or _is_payroll_summary_row(emp_id_str):
            continue

        net_pay = _pv(row, "Net Pay")
        if net_pay == 0:
            continue

        bank_account = _sv(row, "Bank Account Number") or _sv(row, "Bank Account No")
        bank_name    = _sv(row, "Bank Name")
        id_number    = _sv(row, "IC Number") or _sv(row, "Passport Number") or _sv(row, "ID Number")

        gross  = _pv(row, "Gross earnings") or _pv(row, "Gross Salary") or _pv(row, "Gross Pay")
        hrdf   = _pv(row, "HRDF")
        mtd    = _pv(row, "PCB") or _pv(row, "MTD")
        claim  = _pv(row, "Claims Amount") or _pv(row, "Claim")

        # EPF, SOCSO & EIS are computed from the official statutory tables using
        # gross earnings as the contribution wage, rather than trusting whatever
        # the payroll export reports. Category (under/over 60), EPF foreigner
        # status and EIS eligibility are derived from Age and Nationality.
        age         = _pv(row, "Age") or None
        nationality = _sv(row, "Nationality")
        epf_ee,   epf_er   = epf_contribution(gross, age, nationality)
        socso_ee, socso_er = socso_contribution(gross, age)
        eis_ee,   eis_er   = eis_contribution(gross, age, nationality)

        # CTC Hexa = Gross + employer statutory (EPF/EIS/SOCSO) + HRDF + Claims.
        ctc = gross + epf_er + eis_er + socso_er + hrdf + claim

        employees.append({
            "employeeId":    emp_id_str,
            "name":          _sv(row, "Employee Name"),
            "costCentre":    _sv(row, "Department") or _sv(row, "Cost Centre"),
            "grossSalary":   round(gross, 2),
            "netSalary":     round(net_pay, 2),
            "claim":         round(claim, 2),
            "age":           age,
            "nationality":   nationality,
            "epfEmployee":   round(epf_ee, 2),
            "epfEmployer":   round(epf_er, 2),
            "epfBasis":      epf_basis(age, nationality),
            "eisEmployee":   round(eis_ee, 2),
            "eisEmployer":   round(eis_er, 2),
            "socsoEmployee": round(socso_ee, 2),
            "socsoEmployer": round(socso_er, 2),
            "hrdf":          round(hrdf, 2),
            "mtd":           round(mtd, 2),
            "ctcHexa":       round(ctc, 2),
            "bankName":      bank_name,
            "bankAccount":   bank_account,
            "idNumber":      id_number,
        })

    entity_name = company_name or "HSSB Payroll"
    return _json_safe([{
        "sheetName":      entity_name,
        "employees":      employees,
        "totalCTC":       round(sum(e["ctcHexa"] for e in employees), 2),
        "totalNetPay":    round(sum(e["netSalary"] for e in employees), 2),
        "missingColumns": missing_cols,
    }])
