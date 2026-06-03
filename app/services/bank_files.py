import io
import os
import re
import base64
import hashlib
from datetime import datetime, timezone
import httpx
import openpyxl
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from app.config import (
    AIRTABLE_API_KEY, AIRTABLE_BASE_ID, AIRTABLE_TABLE_NAME, BANK_NOTIFY_EMAILS,
)

# ── Fixed values the RCGEN2 macro reads from the workbook's Home/Domestic-Payments
#    sheets. Mirrored here so we can build the bank .txt directly in Python (the
#    macro's own Generate button is Windows-only). If the Maybank corporate
#    registration changes, update these AND the rcgen_template.xlsm Home sheet.
#    Source cells: Home!E5, Home!E6, Home!E7, Domestic Payments!B3, !G2.
RCGEN_CORPORATE_ID   = "MYMHEXAMATI"              # Home!E5  — header field 2
RCGEN_CLIENT_BATCH   = "MYMHEXA1D"               # Home!E6  — header field 3
RCGEN_DEBIT_ACCOUNT  = "514123216966"            # Home!E7  — body field 15
RCGEN_PROC_INDICATOR = "B"                       # DomPay!B3 — header field 5
RCGEN_PRODUCT        = "Domestic Payments (MY)"  # DomPay!G2 — body field 3

# TripleDES key/IV embedded in the RCGEN2 macro (Generate_encryptedMessage). The
# header's encryption token is Base64(3DES-CBC-ZeroPad(filename_run_totalhash)).
_RCGEN_3DES_KEY = b">tlF8adk=35K{dsb"   # 16 bytes → 2-key 3DES (K1,K2,K1)
_RCGEN_3DES_IV  = b"zlrs$5kb"           # 8 bytes
_RCGEN_SEP      = "|"
_RCGEN_EOL      = "\r\n"

# The official Maybank RCGEN2 macro workbook. We fill its "Domestic Payments"
# data sheet (leaving the VBA macro and all lookup sheets intact) so the maker
# can open it and click the workbook's own Generate button to produce a valid
# RCgen_Payment_DP_*.txt — instead of us hand-building that .txt (which the CMS
# portal rejects line-by-line because only the macro emits the exact format).
RCGEN_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "assets", "rcgen_template.xlsm")

MY_BANK_CODES = {
    "maybank": "MBBEMYKL", "maybank islamic": "MBBEMYKL",
    "public bank": "PBBEMYKL", "public bank berhad": "PBBEMYKL",
    "cimb": "CIBBMYKL", "cimb bank": "CIBBMYKL",
    "rhb": "RHBBMYKL", "rhb bank": "RHBBMYKL",
    "hong leong": "HLBBMYKL", "hong leong bank": "HLBBMYKL",
    "ambank": "ARBKMYKL",
    "bank islam": "BIMBMYKL", "bank islam malaysia berhad": "BIMBMYKL",
    "bank muamalat": "BMMBMYKL",
    "hsbc": "HBMBMYKL", "hsbc bank": "HBMBMYKL",
    "ocbc": "OCBCMYKL",
    "standard chartered": "SCBLMYKL",
    "affin": "PHBMMYKL", "affin bank": "PHBMMYKL",
    "alliance bank": "MFBBMYKL",
    "bank rakyat": "BKRMMYKL",
    "bsn": "BSNAMYK1",
}


def bank_name_to_code(name: str) -> str:
    if not name:
        return ""
    return MY_BANK_CODES.get(name.strip().lower(), "")


def _payment_mode(bank_code: str) -> str:
    """IT = Interbank Transfer (Maybank-to-Maybank). IG = IBG (to other banks)."""
    return "IT" if bank_code == "MBBEMYKL" else "IG"


def _strip_spaces_dashes(value: str) -> str:
    """Remove spaces and hyphens (Maybank rejects separators in IC / account
    numbers). e.g. '900101-01-5523' → '900101015523', '1234 5678' → '12345678'."""
    return (value or "").replace(" ", "").replace("-", "").strip()


def _split_name(full: str, max_len: int = 40) -> tuple:
    """Split a beneficiary name into (Name 1, Name 2) for the Maybank file.

    Bank-accepted RCgen files keep the **full name in Name 1** (up to 40 chars)
    and leave Name 2 empty — matching it also helps IBG beneficiary-name checks.
    Only when the full name exceeds ``max_len`` do we overflow trailing words
    into Name 2 (preserving order) until Name 1 fits or one word remains."""
    full = (full or "").strip()
    if len(full) <= max_len:
        return (full, "")
    tokens = full.split()
    if len(tokens) <= 1:
        return (full[:max_len], full[max_len:])
    name1_tokens = tokens[:-1]
    name2_tokens = [tokens[-1]]
    while len(" ".join(name1_tokens)) > max_len and len(name1_tokens) > 1:
        name2_tokens.insert(0, name1_tokens.pop())
    return (" ".join(name1_tokens), " ".join(name2_tokens))


# Exact column headers (row 4) of the Maybank RCGEN2 "Domestic Payments" R3
# template. RCGEN2 reads the header block (rows 1-3), these headers (row 4),
# and data from row 5 — replicated verbatim so the file imports cleanly.
RCMS_DP_HEADERS = [
    "Payment Mode\nIT = INTRABANK\nIG = GIRO\nIM = RENTAS\nAllowed\nValue(IT,IG & IM)\n",
    "Value Date\n[DDMMYYYY]\ne.g.  21102015\n(If start with 0,\nthen add\napostrophe e.g. '0)",
    "Customer \nReference Number\n(If start with 0, then add apostrophe e.g. '0)",
    "Favourite Beneficiary Code",
    "Transaction Amount\n(RM)",
    "Credit Account Number\n(If start with 0, then add\napostrophe e.g. '0)",
    "Beneficiary Name 1\n(Maximum Length is 40)",
    "Beneficiary Name 2\n(Maximum Length is 40)",
    "Beneficiary Name 3\n(Maximum Length is 40)",
    "New NRIC\n(If start with 0,\nthen add\napostrophe e.g. '0)",
    "Old NRIC\n(If start with 0,\nthen add\napostrophe e.g. '0)",
    "Business Registration No\n(If start with 0,\nthen add\napostrophe e.g. '0)",
    "Police/ Army ID/ Passport No\n(If start with 0,\nthen add\napostrophe e.g. '0)",
    "Beneficiary\nBank Code",
    "Email",
    "Advice Detail\n(Maximum Length is 400)\nThis field is mandatory if email exist.",
    "Debit \nDescription",
    "Credit \nDescription",
    "Joint Name (Only applicable for Payment Mode IM)",
    "Joint New ID No (Only applicable for Payment Mode IM)",
    "Joint Old ID No (Only applicable for Payment Mode IM)",
    "Joint Business Reg. No. (Only applicable for Payment Mode IM)",
    "Joint Police/ Army ID/ Passport No. (Only applicable for Payment Mode IM)",
    "Purpose of Transfer (Only applicable for Payment Mode IM) Kindly refer to the list of Purpose of Transfer for RENTAS",
    "Others\xa0 Purpose of Transfer (Only applicable for Payment Mode IM). Free text field (Maximum Length is 35)",
    "Rentas Instruction to Bank (Only applicable for Payment Mode IM)",
    "Charges Borne By\n01 = Applicant\n02 = Beneficiary\n03 = Shared",
] + [f"Email {n}" for n in range(2, 21)]


def _client_initials(client: str) -> str:
    """Client name initials for the payment advice. Multi-word client → initials
    (Bank Negara Malaysia → BNM); single word → the word as-is (Nokia → Nokia)."""
    words = [w for w in re.split(r"[^A-Za-z0-9]+", client or "") if w]
    if not words:
        return ""
    if len(words) == 1:
        return words[0]
    return "".join(w[0] for w in words).upper()


def _advice_detail(client: str, name: str, mmyy: str) -> str:
    """Payment advice format: {ClientInitials}_{First}_{Second}_{MMYY}
    e.g. BNM_Abu_Zharr_0626."""
    parts = [_client_initials(client)] + (name or "").split()[:2] + [mmyy]
    return "_".join(p for p in parts if p)


# Columns written as TEXT so leading zeros and IT/IG are preserved verbatim
# (1-based: Payment Mode, Value Date, Credit Acct, New NRIC, Old NRIC,
# Business Reg, Police/Passport).
_RCMS_TEXT_COLS = {1, 2, 6, 10, 11, 12, 13}


def _dp_row_cells(b, value_date, mmyy, notify_emails, advice_fn=None) -> dict:
    """Build the 'Domestic Payments' sheet cell values (1-indexed column → value)
    for one beneficiary, using the exact RCGEN2 column order (col 1 = Payment Mode
    … col 18 = Credit Description, cols 28/29 = Email 2/3). Single source of truth
    shared by the .xlsm writer (`_write_rcms_dp_row`) and the .txt builder
    (`build_dp_txt`). ``advice_fn(b)`` builds the advice/debit/credit description;
    defaults to the {ClientInitials}_{First}_{Second}_{MMYY} EOR format."""
    advice = advice_fn(b) if advice_fn else _advice_detail(b.get("costCentre", ""), b["name"], mmyy)
    name1, name2 = _split_name(b["name"])
    new_ic, biz_reg, passport = _id_fields(b.get("idNumber", ""), b.get("idType", ""))
    vals = {
        1:  b["paymentMode"],                  # Payment Mode (IT/IG) — text
        2:  value_date,                        # Value Date DDMMYYYY — text
        3:  b["seq"],                          # Customer Reference Number
        4:  b.get("favouriteBeneficiaryCode", ""),  # Favourite Beneficiary/Biller Code (must be registered in Maybank CMS)
        5:  float(b["amount"] or 0),           # Transaction Amount (RM) — number, no comma
        6:  b["accountNumber"],                # Credit Account Number — text
        7:  name1,                             # Beneficiary Name 1 (<=40)
        8:  name2,                             # Beneficiary Name 2 (overflow)
        10: new_ic,                            # New NRIC
        12: biz_reg,                           # Business Registration No
        13: passport,                          # Police/ Army ID/ Passport No
        14: b["bankCode"],                     # Beneficiary Bank Code
        15: notify_emails[0] if notify_emails else "",   # Email
        16: advice,                            # Advice Detail
        17: advice,                            # Debit Description
        18: advice,                            # Credit Description
    }
    if len(notify_emails) > 1:
        vals[28] = notify_emails[1]            # Email 2
    if len(notify_emails) > 2:
        vals[29] = notify_emails[2]            # Email 3
    return vals


def _write_rcms_dp_row(ws, r, b, value_date, mmyy, notify_emails, advice_fn=None):
    """Write a single beneficiary onto row ``r`` of a 'Domestic Payments' sheet."""
    vals = _dp_row_cells(b, value_date, mmyy, notify_emails, advice_fn)
    for col, v in vals.items():
        cell = ws.cell(row=r, column=col, value=v)
        if col in _RCMS_TEXT_COLS:
            cell.number_format = "@"
        elif col == 5:
            cell.number_format = "0.00"


def _fill_rcms_template(beneficiaries, value_date, mmyy, notify_emails, advice_fn=None) -> bytes:
    """Load the official RCGEN2 macro workbook, clear the sample rows from the
    'Domestic Payments' sheet, and write our beneficiaries into it — preserving
    the VBA macro and every lookup sheet. The maker opens the returned .xlsm and
    clicks the workbook's Generate button to emit a valid RCgen .txt.

    Only beneficiaries with a bank account are written (the macro would reject a
    blank Credit Account Number), matching the rows that belong in the payment."""
    wb = openpyxl.load_workbook(RCGEN_TEMPLATE_PATH, keep_vba=True)
    ws = wb["Domestic Payments"]
    # Rows 1-4 are the template's header block / column titles — leave intact.
    # Drop any pre-existing sample data (row 5 onward) before writing ours.
    if ws.max_row >= 5:
        ws.delete_rows(5, ws.max_row - 4)
    r = 5
    for b in beneficiaries:
        if not b["accountNumber"]:
            continue
        _write_rcms_dp_row(ws, r, b, value_date, mmyy, notify_emails, advice_fn)
        r += 1
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  Direct .txt generation — a faithful Python port of the RCGEN2 macro's
#  File_Single_DomPay so the maker can download the bank .txt without Excel.
#  Field-by-field equivalent of the VBA; validate byte-for-byte against a known
#  macro-generated file before relying on it.
# ─────────────────────────────────────────────────────────────────────────────
def _lc(val, n: int) -> str:
    """LengthCheck(): stringify and truncate to n chars (macro left-aligns, no pad)."""
    s = "" if val is None else str(val)
    return s[:n]


def _rcgen_token(filename: str, run_number, total_hash: int) -> str:
    """Header field 6 = Base64(3DES-CBC, Zeros-pad) of '<filename>_<run>_<hash>'.
    Reproduces Generate_encryptedMessage from the RCGEN2 macro exactly."""
    plain = f"{filename}_{run_number}_{total_hash}".encode("utf-8")
    data = plain + b"\x00" * ((-len(plain)) % 8)           # .NET PaddingMode.Zeros
    enc = Cipher(algorithms.TripleDES(_RCGEN_3DES_KEY), modes.CBC(_RCGEN_3DES_IV)).encryptor()
    return base64.b64encode(enc.update(data) + enc.finalize()).decode()


def _row_hash(amount: float, account_no: str, count: int) -> int:
    """Per-row hash, matching the macro: amount-mod-2000 + a digit/ASCII sum of the
    last 6 credit-account characters, each offset by the running row count."""
    amt = float(amount) * 100
    s_hash = (amt - (amt // 2000) * 2000) + count
    accno = account_no or ""
    if len(accno) < 6:
        accno = accno.rjust(6, "0")
    last6 = accno[-6:]
    acc_val = sum(int(ch) if ch.isdigit() else ord(ch) for ch in last6)
    acc_val = acc_val * 2 + count
    return int(round(s_hash)) + int(round(acc_val))


def _join_sep_terminated(fields: list) -> str:
    """Every field followed by '|', then EOL — 'f1|f2|...|fn|\\r\\n' (body record)."""
    return "".join(f + _RCGEN_SEP for f in fields) + _RCGEN_EOL


def _join_open_last(fields: list) -> str:
    """All but the last field followed by '|', last field bare, then EOL —
    'f1|...|f(n-1)|fn\\r\\n' (header / advice / trailer records)."""
    return "".join(f + _RCGEN_SEP for f in fields[:-1]) + (fields[-1] if fields else "") + _RCGEN_EOL


def _dp_header_record(token: str) -> str:
    fields = ["00",
              _lc(RCGEN_CORPORATE_ID.upper(), 30),
              _lc(RCGEN_CLIENT_BATCH, 30),
              "",                       # 4 Account Payees Only
              RCGEN_PROC_INDICATOR,     # 5 Processing Indicator
              token]                    # 6 encryption RCGEN
    fields += [""] * 23                 # 7–29 fillers
    return _join_open_last(fields)      # 29 fields


def _dp_body_record(c: dict, amount: float) -> str:
    amt2 = f"{float(amount):.2f}"
    b = [""] * 167
    b[0]   = "01"
    b[1]   = _lc(str(c.get(1, "")).upper(), 2)   # 2 Payment Mode
    b[2]   = _lc(RCGEN_PRODUCT, 50)              # 3 Product
    b[4]   = str(c.get(2, ""))                   # 5 Value Date
    b[7]   = _lc(c.get(3, ""), 30)               # 8 Customer Reference Number
    b[9]   = _lc(c.get(17, ""), 55)              # 10 Debit Description
    b[10]  = "MYR"                               # 11 Transaction Currency
    b[11]  = amt2                                # 12 Transaction Amount
    b[12]  = "Y"                                 # 13 In Debit Account Currency
    b[13]  = "MYR"                               # 14 Debiting Currency
    b[14]  = _lc(RCGEN_DEBIT_ACCOUNT, 20)        # 15 Debiting Account Number
    b[15]  = _lc(c.get(6, ""), 35)               # 16 Credit Account Number
    b[16]  = _lc(c.get(4, ""), 15)               # 17 Favourite Beneficiary/Biller Code
    b[18]  = "Y"                                 # 19 Resident Indicator
    b[19]  = _lc(c.get(7, ""), 40)               # 20 Beneficiary Name 1
    b[20]  = _lc(c.get(8, ""), 40)               # 21 Beneficiary Name 2
    b[21]  = _lc(c.get(9, ""), 40)               # 22 Beneficiary Name 3
    b[24]  = _lc(c.get(10, ""), 20)              # 25 New ID No
    b[25]  = _lc(c.get(11, ""), 20)              # 26 Old ID No
    b[26]  = _lc(c.get(12, ""), 20)              # 27 Business Registration No
    b[27]  = _lc(c.get(13, ""), 20)              # 28 Police/Army ID/Passport No
    b[36]  = str(c.get(14, ""))                  # 37 Beneficiary Bank Code
    b[102] = _lc(c.get(18, ""), 55)              # 103 Credit Description
    b[109] = "01"                                # 110 Charges Borne By
    b[110] = _lc(c.get(24, ""), 5)               # 111 Purpose of Transfer
    b[160] = _lc(c.get(19, ""), 32)              # 161 Joint Name
    b[161] = _lc(c.get(20, ""), 20)              # 162 Joint New ID No
    b[162] = _lc(c.get(21, ""), 20)              # 163 Joint Old ID No
    b[163] = _lc(c.get(22, ""), 20)              # 164 Joint Business Reg. No.
    b[164] = _lc(c.get(23, ""), 20)              # 165 Joint Police/Army ID/Passport No.
    b[165] = _lc(c.get(25, ""), 35)              # 166 Others Purpose of Transfer
    b[166] = _lc(c.get(26, ""), 66)              # 167 Rentas Instruction to Bank
    b += [""] * 168                              # 168–335 fillers
    b += [""]                                    # 336 Transaction Return Status
    return _join_sep_terminated(b)               # 336 fields, trailing '|'


def _dp_advice_record(c: dict, amount: float) -> str:
    amt2 = f"{float(amount):.2f}"
    a = [""] * 40
    a[0] = "02"
    a[1] = "PA"
    a[2] = _lc(c.get(3, ""), 30)                 # 3 Customer Reference Number
    a[3] = _lc(c.get(15, ""), 80)                # 4 Email
    a[6] = _lc(c.get(16, ""), 400)               # 7 Advice Detail
    a[13] = amt2                                 # 14 Payment Advice Amount
    for i, col in enumerate(range(28, 47)):      # 21–39 Email 2–20 (cols 28–46)
        a[20 + i] = _lc(c.get(col, ""), 80)
    return _join_open_last(a)                     # 40 fields


def _dp_trailer_record(count: int, trans_amount: float, total_hash: int) -> str:
    t = ["99", str(count), f"{float(trans_amount):.2f}", str(total_hash)]
    t += [""] * 26                                # 5–30 fillers
    return _join_open_last(t)                      # 30 fields


def build_dp_txt(beneficiaries, value_date, mmyy, run_number, notify_emails,
                 advice_fn=None, now=None) -> tuple:
    """Build the Maybank RCGEN Domestic-Payments .txt directly (no Excel macro).
    Returns (filename, text). Only beneficiaries with a credit account and a
    positive amount are included — matching the rows the .xlsm path writes and
    the macro's own amount>0 filter. ``run_number`` is the RCGEN running number
    (must be managed by the caller, like the macro's runningnumber sheet)."""
    now = now or datetime.now()
    filename = "RCgen_Payment_DP_" + now.strftime("%d%m%Y%H%M%S") + ".txt"

    bodies, count, total_hash, trans_amount = [], 0, 0, 0.0
    for b in beneficiaries:
        if not b.get("accountNumber"):
            continue
        amount = float(b.get("amount") or 0)
        if amount <= 0:
            continue
        count += 1
        c = _dp_row_cells(b, value_date, mmyy, notify_emails, advice_fn)
        total_hash += _row_hash(amount, b["accountNumber"], count)
        trans_amount += amount
        rec = _dp_body_record(c, amount)
        if str(c.get(15, "")) != "":
            rec += _dp_advice_record(c, amount)
        bodies.append(rec)

    if count == 0:
        return filename, ""   # nothing to pay → no file (macro Exits)

    token = _rcgen_token(filename, run_number, total_hash)
    text = _dp_header_record(token) + "".join(bodies) + _dp_trailer_record(count, trans_amount, total_hash)
    return filename, text


def _id_fields(id_number: str, id_type: str = "") -> tuple:
    """Return (new_ic, biz_reg, passport) for fields 25, 27, 28 of the 01 record.

    When the Airtable ``ID Type`` is known it is authoritative — NRIC → New IC No
    (field 25), Passport → Police/Army ID/Passport No (field 28). Otherwise fall
    back to inferring from the number's format. Spaces/hyphens are stripped so a
    dashed NRIC (e.g. 900101-01-5523) is recognised as a 12-digit IC."""
    id_str = _strip_spaces_dashes(id_number)
    if not id_str:
        return ("", "", "")

    t = (id_type or "").strip().lower()
    if t == "nric":
        return (id_str, "", "")
    if t == "passport":
        return ("", "", id_str)

    if id_str.isdigit() and len(id_str) == 12:
        return (id_str, "", "")   # Malaysian NRIC → New IC No (field 25)
    # Count leading alpha chars to distinguish passport from company reg
    lead = 0
    for c in id_str:
        if c.isalpha():
            lead += 1
        else:
            break
    if lead == 1:
        return ("", "", id_str)   # Single-letter prefix → Passport/Police/Army (field 28)
    return ("", id_str, "")       # 0 or 2+ leading letters → Business Reg No (field 27)


async def fetch_airtable_consultants() -> list[dict]:
    if not AIRTABLE_API_KEY or not AIRTABLE_BASE_ID or not AIRTABLE_TABLE_NAME:
        return []
    records = []
    offset = None
    async with httpx.AsyncClient(timeout=30) as client:
        while True:
            params = {
                "pageSize": 100,
                "cellFormat": "string",
                "timeZone": "Asia/Kuala_Lumpur",
                "userLocale": "en-MY",
            }
            if offset:
                params["offset"] = offset
            resp = await client.get(
                f"https://api.airtable.com/v0/{AIRTABLE_BASE_ID}/{AIRTABLE_TABLE_NAME}",
                headers={"Authorization": f"Bearer {AIRTABLE_API_KEY}"},
                params=params,
            )
            if not resp.is_success:
                break
            data = resp.json()
            for r in data.get("records", []):
                f = r.get("fields", {})
                records.append({
                    "employeeNumber": str(f.get("Employee Number", "")).strip(),
                    "employeeId": str(f.get("Employee ID", "")).strip(),
                    "name": str(f.get("Full Legal Name", "")).strip(),
                    "bankName": str(f.get("Bank Name", "")).strip(),
                    "accountNo": str(f.get("Bank Account Number", "")).strip(),
                    "idNumber": str(f.get("ID Number", "")).strip(),
                    "idType": str(f.get("ID Type", "") or "").strip(),
                    "favouriteBeneficiaryCode": str(f.get("Favourite Beneficiary Code", "") or "").strip(),
                    "nationality": str(f.get("Nationality", "") or "").strip(),
                    "contractType": str(f.get("Contract Type", "") or "").strip(),
                    "epfScheme": str(f.get("EPF Scheme", "") or "").strip(),
                    "epfNumber":   str(f.get("EPF Number", "") or "").strip(),
                    "socsoNumber": str(f.get("SOCSO Number", "") or "").strip(),
                    "taxRefNumber":str(f.get("Tax Identification Number", "") or "").strip(),
                })
            offset = data.get("offset")
            if not offset:
                break
    return records


def match_consultant(emp: dict, airtable_list: list[dict]):
    by_num = next(
        (a for a in airtable_list if a["employeeNumber"] == emp.get("employeeId") or a["employeeId"] == emp.get("employeeId")),
        None,
    )
    if by_num:
        return by_num
    emp_lower = emp.get("name", "").lower()
    return next(
        (a for a in airtable_list if a["name"].lower() == emp_lower or emp_lower in a["name"].lower() or a["name"].lower() in emp_lower),
        None,
    )


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def next_rcgen_run_number(db, key: str = "rcgen_dp") -> int:
    """Atomically increment and return the persistent RCGEN running number — the
    DB equivalent of the macro's runningnumber!A2 (the bank uses it for anti-
    replay). Backed by a Postgres function `next_counter(p_key text)` over an
    `app_counters` table. Raises if that object is missing so the caller can
    decide to skip .txt generation rather than reuse a number."""
    resp = db.rpc("next_counter", {"p_key": key}).execute()
    val = resp.data
    if isinstance(val, list):          # some PostgREST configs wrap scalars
        val = val[0]
    if isinstance(val, dict):
        val = next(iter(val.values()))
    return int(val)


async def generate_and_store_bank_files(kase: dict, db, triggered_by: str) -> dict:
    entities = (kase.get("parsed_data") or {}).get("entities", [])
    check = kase.get("check_data") or {}
    now = datetime.now(timezone.utc).isoformat()

    payment_date_str = kase.get("payment_date") or now[:10]
    yr, mo, dy = payment_date_str.split("-")
    value_date = f"{dy}{mo}{yr}"
    mmyy = f"{mo}{yr[2:]}"

    airtable_list: list[dict] = []
    try:
        airtable_list = await fetch_airtable_consultants()
    except Exception:
        pass

    notify_emails = BANK_NOTIFY_EMAILS

    beneficiaries = []
    excluded_no_fav = []   # consultants with no Favourite Beneficiary Code — skipped + flagged
    seq_ref = 100
    for ent in entities:
        for emp in ent.get("employees", []):
            matched = match_consultant(emp, airtable_list)
            bank_code = bank_name_to_code(matched["bankName"] if matched else "")

            # Favourite Beneficiary/Biller Code: the CSI value wins, else the
            # consultant-DB (Airtable) value. The bank only accepts a code that is
            # registered as a favourite in Maybank CMS, so a consultant without one
            # is EXCLUDED from the file (others continue) and flagged for follow-up.
            fav_code = (emp.get("favouriteBeneficiaryCode") or "").strip() \
                or (matched.get("favouriteBeneficiaryCode", "").strip() if matched else "")
            if not fav_code:
                excluded_no_fav.append({"name": (matched["name"] if matched else emp.get("name", "")),
                                        "employeeId": emp.get("employeeId", ""),
                                        "entity": ent["sheetName"]})
                continue

            beneficiaries.append({
                "seq": seq_ref,
                "employeeId": emp["employeeId"],
                "employeeCode": matched["employeeNumber"] if matched else emp.get("employeeId", ""),
                "favouriteBeneficiaryCode": fav_code,
                "name": matched["name"] if matched else emp["name"],
                "costCentre": emp.get("costCentre", ""),
                "amount": emp.get("netSalary", 0),
                "accountNumber": _strip_spaces_dashes(matched["accountNo"] if matched else ""),
                "bankName": matched["bankName"] if matched else "",
                "bankCode": bank_code,
                "paymentMode": _payment_mode(bank_code),
                "email": notify_emails[0] if notify_emails else "",
                "idNumber": matched["idNumber"] if matched else emp.get("idNumber", ""),
                "idType": (matched.get("idType") if matched else "") or emp.get("idType", ""),
                "advicePrefix": (matched["name"] if matched else emp["name"]).replace(" ", "_"),
                "entity": ent["sheetName"],
                "matched": matched is not None,
            })
            seq_ref += 1

    # ── Fill the official RCGEN2 macro workbook ('Domestic Payments' sheet).
    #    The maker opens this .xlsm and clicks its Generate button to emit the
    #    valid RCgen .txt — we no longer hand-build that .txt ourselves. ──
    xlsx_bytes = _fill_rcms_template(beneficiaries, value_date, mmyy, notify_emails)
    xlsx_hash = _sha256(xlsx_bytes)
    xlsx_name = f"RCMS_Payment_DP_{kase['reference']}_{value_date}.xlsm"

    missing = [{"name": b["name"], "employeeId": b["employeeId"]} for b in beneficiaries if not b["matched"]]
    existing_check = dict(kase.get("check_data") or {})
    existing_check["missingBankAccounts"] = missing
    existing_check["excludedNoFavourite"] = excluded_no_fav

    # Also build the bank .txt directly (no Excel macro). Best-effort: if the
    # running-number counter isn't provisioned yet, skip silently — the .xlsm
    # path is unaffected and remains the primary download.
    try:
        run_number = next_rcgen_run_number(db)
        txt_name, txt_body = build_dp_txt(beneficiaries, value_date, mmyy, run_number, notify_emails)
        existing_check["bankTxt"] = {"name": txt_name, "runNumber": run_number,
                                     "data": base64.b64encode(txt_body.encode("utf-8")).decode()}
    except Exception as e:
        existing_check["bankTxt"] = None
        existing_check["bankTxtError"] = str(e)[:200]

    db.from_("payroll_cases").update({
        "status":                 "bank_file_generated",
        "bank_file_name":         xlsx_name,
        "bank_file_hash":         xlsx_hash,
        "bank_file_data":         base64.b64encode(xlsx_bytes).decode(),
        "bank_file_generated_at": now,
        "bank_file_triggered_by": triggered_by,
        "bank_receipt_name":      None,
        "bank_receipt_data":      None,
        "check_data":             existing_check,
    }).eq("id", kase["id"]).execute()

    matched_count = sum(1 for b in beneficiaries if b["matched"])
    return {
        "xlsxName": xlsx_name,
        "xlsxBytes": xlsx_bytes,
        "matched": matched_count,
        "total": len(beneficiaries),
        "missing": missing,
        "excludedNoFavourite": excluded_no_fav,
    }


async def generate_and_store_bank_files_payroll(kase: dict, db, triggered_by: str) -> dict:
    """
    Generate bank files for PAYROLL (internal employees).
    Bank details come directly from the parsed payroll file — no Airtable lookup needed.
    Only net salary payments are included; statutory contributions (EPF/SOCSO/HRDF/PCB)
    are paid separately through government portals.
    """
    entities = (kase.get("parsed_data") or {}).get("entities", [])
    check = kase.get("check_data") or {}
    now = datetime.now(timezone.utc).isoformat()

    payment_date_str = kase.get("payment_date") or now[:10]
    yr, mo, dy = payment_date_str.split("-")
    value_date = f"{dy}{mo}{yr}"
    mmyy = f"{mo}{yr[2:]}"

    notify_emails = BANK_NOTIFY_EMAILS

    beneficiaries = []
    seq_ref = 100
    for ent in entities:
        for emp in ent.get("employees", []):
            bank_name    = emp.get("bankName", "")
            bank_account = _strip_spaces_dashes(emp.get("bankAccount", ""))
            bank_code    = bank_name_to_code(bank_name)
            has_bank     = bool(bank_account)
            name         = emp.get("name", emp.get("employeeId", ""))
            beneficiaries.append({
                "seq":          seq_ref,
                "employeeId":   emp["employeeId"],
                "employeeCode": emp.get("employeeId", ""),
                "name":         name,
                "costCentre":   emp.get("costCentre", ""),
                "amount":       emp.get("netSalary", 0),
                "accountNumber": bank_account,
                "bankName":     bank_name,
                "bankCode":     bank_code,
                "paymentMode":  _payment_mode(bank_code),
                "email":        notify_emails[0] if notify_emails else "",
                "idNumber":     emp.get("idNumber", ""),
                "idType":       emp.get("idType", ""),
                "advicePrefix": name.replace(" ", "_"),
                "entity":       ent["sheetName"],
                "matched":      has_bank,
            })
            seq_ref += 1

    # ── Fill the official RCGEN2 macro workbook ('Domestic Payments' sheet).
    #    Payroll advice format is {FullName_Underscored}_{MMYY}. ──
    xlsx_bytes = _fill_rcms_template(
        beneficiaries, value_date, mmyy, notify_emails,
        advice_fn=lambda b: f"{b['advicePrefix']}_{mmyy}",
    )
    xlsx_hash = _sha256(xlsx_bytes)
    xlsx_name = f"RCMS_Payment_DP_{kase['reference']}_{value_date}.xlsm"

    missing = [{"name": b["name"], "employeeId": b["employeeId"]} for b in beneficiaries if not b["matched"]]
    existing_check = dict(kase.get("check_data") or {})
    existing_check["missingBankAccounts"] = missing

    db.from_("payroll_cases").update({
        "status":                   "bank_file_generated",
        "bank_file_name":           xlsx_name,
        "bank_file_hash":           xlsx_hash,
        "bank_file_data":           base64.b64encode(xlsx_bytes).decode(),
        "bank_file_generated_at":   now,
        "bank_file_triggered_by":   triggered_by,
        "bank_receipt_name":        None,
        "bank_receipt_data":        None,
        "check_data":               existing_check,
    }).eq("id", kase["id"]).execute()

    matched_count = sum(1 for b in beneficiaries if b["matched"])
    return {
        "xlsxName":  xlsx_name,
        "xlsxBytes": xlsx_bytes,
        "matched":   matched_count,
        "total":     len(beneficiaries),
        "missing":   missing,
    }
