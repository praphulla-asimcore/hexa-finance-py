"""Import pipeline for consultant_master, sourced from the Talenox HR export
("Profiles-employees-export-Hexamatics .xlsx" or any later refresh of it).

Replaces Airtable as a bank-detail data source (see app/services/bank_files.py
fetch_consultant_master / build_consultant_list). Designed to be re-run
repeatedly as the user updates the source file -- e.g. once Favourite
Beneficiary Codes get filled in -- not a one-off script.

Design notes
------------
* The export's employee_id is a mix of already-assigned APEX ids (HEX-0030)
  and entity-native Talenox ids (HC-24, 358, plain numbers) for people not yet
  onboarded into APEX, and is NOT guaranteed unique even within one sheet (a
  real duplicate was found: "HC-01" used for two different people in the HEDU
  sheet). find_duplicate_keys() catches this -- those rows are excluded from
  upsert and reported, never silently merged under a shared key.
* resolve_apex_id() never guesses: an ambiguous IC or name match (more than
  one candidate) resolves to None rather than picking one, matching the
  "fail loud" philosophy of bank_files.match_consultant (the same one that
  caught the Azeean/Nazarul mis-payment defect).
* sync_fav_codes_to_overrides() only ever writes favourite_beneficiary_code
  onto an EXISTING consultant_bank_overrides row, or inserts a fresh row
  populated with this same import's own bank details. It deliberately never
  inserts a fav-code-only row with blank bank fields -- consultant_bank_overrides
  wins bank-file precedence as a whole record (see build_consultant_list), so a
  partial row would shadow good account data from consultant_master/HexaFlow
  with a blank account number.
"""
import re
from datetime import date, datetime

import openpyxl

from app.config import DATABASE_URL

SHEETS = ("HSSB", "HCSSB", "HEDU", "DATACRATS")

_HEX_ID_RE = re.compile(r"^HEX-\d+$", re.IGNORECASE)


def _norm_name(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (s or "").lower()).strip()


def _strip(v) -> str:
    return str(v).strip() if v is not None else ""


def _infer_ic(ssn: str, passport_number: str) -> tuple:
    """Return (ic_number, ic_type). A 12-digit numeric ssn is a Malaysian
    NRIC; otherwise fall back to the passport_number column."""
    ssn_digits = re.sub(r"[\s-]", "", ssn or "")
    if ssn_digits.isdigit() and len(ssn_digits) == 12:
        return ssn_digits, "NRIC"
    if ssn:
        return ssn.strip(), "Passport"
    if passport_number:
        return passport_number.strip(), "Passport"
    return "", ""


def _parse_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError:
            continue
    return None


def parse_workbook(path: str) -> list[dict]:
    """Parse every known entity sheet into flat consultant rows.

    Talenox's export has 3 header-ish rows before data: row 1 is machine
    column names (what we key off), row 2 is an instructional text blob, row
    3 repeats the human-readable labels. Rather than hardcoding a start row,
    each data row is validated by its employee_id cell being non-blank and
    not one of those header artifacts -- robust to a sheet having a slightly
    different row count.
    """
    from app.services.bank_files import bank_name_to_code

    wb = openpyxl.load_workbook(path, data_only=True)
    rows: list[dict] = []
    for sheet_name in SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h}

        fav_col = next((h for h in idx if "favourite" in h.lower() or "beneficiary" in h.lower()), None)

        for row in ws.iter_rows(min_row=2, values_only=True):
            def get(col):
                i = idx.get(col)
                return row[i] if i is not None and i < len(row) else None

            eid = _strip(get("employee_id"))
            if not eid or eid.lower() == "employee id" or "hi there" in eid.lower():
                continue

            ssn = _strip(get("ssn"))
            passport = _strip(get("passport_number"))
            ic_number, ic_type = _infer_ic(ssn, passport)
            bank_name = _strip(get("bank_account_attributes.bank_type"))

            rows.append({
                "entity": sheet_name,
                "employee_id": eid,
                "consultant_name": " ".join(
                    p for p in [_strip(get("first_name")), _strip(get("last_name"))] if p
                ) or _strip(get("identification_full_name")),
                "ic_number": ic_number,
                "ic_type": ic_type,
                "nationality": _strip(get("nationality")),
                "bank_name": bank_name,
                "bank_code": bank_name_to_code(bank_name),
                "bank_account_name": _strip(get("bank_account_attributes.account_name")),
                "bank_account_number": _strip(get("bank_account_attributes.number")).replace(" ", ""),
                "epf_number": _strip(get("additional_my_employee_info_attributes.epf_number")),
                "tin_number": _strip(get("pcb_number")),
                "resign_date": _parse_date(get("resign_date")),
                "fav_code": _strip(get(fav_col)) if fav_col else "",
            })
    return rows


def find_duplicate_keys(rows: list[dict]) -> dict:
    """Group rows by (entity, employee_id); return only groups that are a
    genuine conflict -- more than one distinct (normalised) name sharing the
    same key, e.g. the HC-01/HEDU case. Repeats of the identical person
    (same name) under one key are not a conflict."""
    groups: dict = {}
    for r in rows:
        key = (r["entity"], r["employee_id"])
        groups.setdefault(key, []).append(r)

    conflicts = {}
    for key, group in groups.items():
        names = {_norm_name(r["consultant_name"]) for r in group if r["consultant_name"]}
        if len(names) > 1:
            conflicts[key] = group
    return conflicts


def build_ic_index(db) -> dict:
    """ic_number -> apex_employee_id, sourced from consultant_directory
    (HexaFlow-standing, keyed by APEX employee_id already). Ambiguous ICs
    (shared by more than one employee_id) are dropped rather than guessed."""
    try:
        rows = db.from_("consultant_directory").select("employee_id,id_number").execute().data or []
    except Exception:
        rows = []
    by_ic: dict = {}
    for r in rows:
        ic = _strip(r.get("id_number"))
        eid = _strip(r.get("employee_id"))
        if not ic or not eid:
            continue
        by_ic.setdefault(ic, set()).add(eid)
    return {ic: next(iter(eids)) for ic, eids in by_ic.items() if len(eids) == 1}


def build_name_index(db) -> dict:
    """normalised name -> set(apex_employee_id), scanned from every historical
    payroll_cases.parsed_data entity/employee -- same approach already proven
    in scripts/import_beneficiary_tracker_v1.py. Only HEX-xxxx ids are
    indexed, since that's the id format CSI/payroll rows carry."""
    try:
        rows = db.from_("payroll_cases").select("parsed_data").execute().data or []
    except Exception:
        rows = []
    by_name: dict = {}
    for r in rows:
        parsed = r.get("parsed_data") or {}
        for ent in parsed.get("entities", []):
            for emp in ent.get("employees", []):
                eid = _strip(emp.get("employeeId"))
                name = _norm_name(emp.get("name", ""))
                if name and eid and _HEX_ID_RE.match(eid):
                    by_name.setdefault(name, set()).add(eid.upper())
    return by_name


def resolve_apex_id(row: dict, ic_index: dict, name_index: dict):
    """Never guesses: returns the resolved HEX-xxxx id, or None when the
    employee is genuinely new to APEX or the match would be ambiguous."""
    eid = row["employee_id"]
    if _HEX_ID_RE.match(eid):
        return eid.upper()

    ic = row.get("ic_number")
    if ic and ic in ic_index:
        return ic_index[ic]

    name = _norm_name(row.get("consultant_name", ""))
    candidates = name_index.get(name)
    if candidates and len(candidates) == 1:
        return next(iter(candidates))

    return None


def upsert_consultant_master(rows: list[dict]) -> dict:
    """ON CONFLICT (entity, employee_id) DO UPDATE. Uses DATABASE_URL via
    psycopg directly, like the other ON CONFLICT upserts in this codebase
    (app/services/hexaflow_finance_profiles.py, bank_files.next_rcgen_run_number)."""
    import psycopg
    from psycopg.rows import dict_row

    inserted, updated = 0, 0
    with psycopg.connect(DATABASE_URL, autocommit=True, row_factory=dict_row) as conn:
        conn.prepare_threshold = None
        for r in rows:
            result = conn.execute(
                """
                INSERT INTO consultant_master
                    (entity, employee_id, apex_employee_id, consultant_name,
                     ic_number, ic_type, nationality,
                     bank_name, bank_code, bank_account_name, bank_account_number,
                     epf_number, tin_number, resign_date, imported_at)
                VALUES (%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,%s, %s,%s,%s, now())
                ON CONFLICT (entity, employee_id) DO UPDATE SET
                    apex_employee_id     = EXCLUDED.apex_employee_id,
                    consultant_name      = EXCLUDED.consultant_name,
                    ic_number            = EXCLUDED.ic_number,
                    ic_type              = EXCLUDED.ic_type,
                    nationality          = EXCLUDED.nationality,
                    bank_name            = EXCLUDED.bank_name,
                    bank_code            = EXCLUDED.bank_code,
                    bank_account_name    = EXCLUDED.bank_account_name,
                    bank_account_number  = EXCLUDED.bank_account_number,
                    epf_number           = EXCLUDED.epf_number,
                    tin_number           = EXCLUDED.tin_number,
                    resign_date          = EXCLUDED.resign_date,
                    imported_at          = now()
                RETURNING (xmax = 0) AS inserted
                """,
                [r["entity"], r["employee_id"], r.get("apex_employee_id"), r["consultant_name"],
                 r["ic_number"], r["ic_type"], r["nationality"],
                 r["bank_name"], r["bank_code"], r["bank_account_name"], r["bank_account_number"],
                 r["epf_number"], r["tin_number"], r["resign_date"]],
            )
            if result.fetchone()["inserted"]:
                inserted += 1
            else:
                updated += 1
    return {"inserted": inserted, "updated": updated}


def sync_fav_codes_to_overrides(rows: list[dict], triggered_by: str = "consultant_master_import") -> dict:
    """For rows with both a resolved apex_employee_id and a non-empty
    fav_code, upsert favourite_beneficiary_code into consultant_bank_overrides.
    An existing row only gets that one field updated (bank details there stay
    independently curated); a brand new row is populated with this import's
    own bank details too, so it's never a fav-code-only partial record that
    would shadow good account data in bank-file precedence."""
    import psycopg
    from datetime import timezone

    candidates = [r for r in rows if r.get("apex_employee_id") and r.get("fav_code")]
    if not candidates:
        return {"updated": 0, "inserted": 0}

    now = datetime.now(timezone.utc)
    updated, inserted = 0, 0
    with psycopg.connect(DATABASE_URL, autocommit=True) as conn:
        conn.prepare_threshold = None
        for r in candidates:
            existing = conn.execute(
                "SELECT employee_id FROM consultant_bank_overrides WHERE employee_id = %s",
                [r["apex_employee_id"]],
            ).fetchone()
            if existing:
                conn.execute(
                    """
                    UPDATE consultant_bank_overrides
                    SET favourite_beneficiary_code = %s, updated_at = %s, updated_by = %s
                    WHERE employee_id = %s
                    """,
                    [r["fav_code"], now, triggered_by, r["apex_employee_id"]],
                )
                updated += 1
            else:
                conn.execute(
                    """
                    INSERT INTO consultant_bank_overrides
                        (employee_id, consultant_name, favourite_beneficiary_code,
                         bank_account_number, bank_code, bank_name,
                         source, updated_at, updated_by)
                    VALUES (%s,%s,%s, %s,%s,%s, %s,%s,%s)
                    """,
                    [r["apex_employee_id"], r["consultant_name"], r["fav_code"],
                     r["bank_account_number"], r["bank_code"], r["bank_name"],
                     "CONSULTANT_MASTER_IMPORT", now, triggered_by],
                )
                inserted += 1
    return {"updated": updated, "inserted": inserted}


def fetch_missing_fav_code_rows(db) -> list[dict]:
    """consultant_master rows resolved to an apex_employee_id that don't yet
    have a non-blank Favourite Beneficiary Code in consultant_bank_overrides
    -- the definitive "still need to register + capture a fav code" list,
    used by the /admin/missing-fav-codes page. Collapses to one row per
    apex_employee_id (a consultant can appear under more than one entity
    sheet)."""
    try:
        master_rows = db.from_("consultant_master").select("*").execute().data or []
        override_rows = db.from_("consultant_bank_overrides").select(
            "employee_id,favourite_beneficiary_code").execute().data or []
    except Exception:
        return []
    has_fav_code = {
        r["employee_id"] for r in override_rows
        if (r.get("favourite_beneficiary_code") or "").strip()
    }
    seen: set = set()
    out = []
    for r in master_rows:
        apex_id = r.get("apex_employee_id")
        if not apex_id or apex_id in has_fav_code or apex_id in seen:
            continue
        seen.add(apex_id)
        out.append(r)
    out.sort(key=lambda r: r.get("consultant_name") or "")
    return out


def set_fav_code_for_consultant(apex_employee_id: str, fav_code: str, triggered_by: str) -> bool:
    """Single-consultant version of sync_fav_codes_to_overrides, for the
    /admin/missing-fav-codes page: looks up this consultant's own bank details
    in consultant_master so the same complete-record guarantee applies (never
    a fav-code-only row that would shadow a good account number in bank-file
    precedence). Returns False if the apex id isn't in consultant_master."""
    import psycopg
    from psycopg.rows import dict_row

    with psycopg.connect(DATABASE_URL, row_factory=dict_row) as conn:
        conn.prepare_threshold = None
        row = conn.execute(
            "SELECT consultant_name, bank_name, bank_code, bank_account_number "
            "FROM consultant_master WHERE apex_employee_id = %s LIMIT 1",
            [apex_employee_id],
        ).fetchone()
    if not row:
        return False
    sync_fav_codes_to_overrides(
        [{**row, "apex_employee_id": apex_employee_id, "fav_code": fav_code}],
        triggered_by,
    )
    return True


def run_import(path: str, db) -> dict:
    """Orchestrates the full pipeline; returns a report dict for the CLI to print."""
    rows = parse_workbook(path)
    duplicates = find_duplicate_keys(rows)
    dup_keys = set(duplicates.keys())
    clean_rows = [r for r in rows if (r["entity"], r["employee_id"]) not in dup_keys]

    ic_index = build_ic_index(db)
    name_index = build_name_index(db)
    for r in clean_rows:
        r["apex_employee_id"] = resolve_apex_id(r, ic_index, name_index)

    master_result = upsert_consultant_master(clean_rows)
    fav_result = sync_fav_codes_to_overrides(clean_rows)

    return {
        "total_rows": len(rows),
        "duplicates_skipped": duplicates,
        "master_inserted": master_result["inserted"],
        "master_updated": master_result["updated"],
        "unresolved_apex_id": sum(1 for r in clean_rows if not r["apex_employee_id"]),
        "fav_codes_inserted": fav_result["inserted"],
        "fav_codes_updated": fav_result["updated"],
    }
